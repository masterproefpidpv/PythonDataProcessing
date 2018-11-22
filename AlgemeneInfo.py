import openpyxl

class AlgemeneInfo():
    def datasheet(workbook, sheetNames, huidigUur, pad):
        wb = workbook
        dataEx = openpyxl.load_workbook(pad + 'data-exchange' + huidigUur + '.xlsx')
        data = dataEx['data-exchange']
        generalSheet = wb['General']

        # uren berekenen
        vorigUur = 0
        uren = [0]
        for i in range(20, generalSheet.max_row, 1):
            if generalSheet.cell(row=i, column=1).value == None:
                break
            else:
                date1 = generalSheet.cell(row=i, column=1).value
                date2 = generalSheet.cell(row=i - 1, column=1).value
                timeOut = generalSheet.cell(row=i, column=2).value
                timeIn = generalSheet.cell(row=i - 1, column=3).value

                date1 = date1.replace(hour=timeOut.hour, minute=timeOut.minute)
                date2 = date2.replace(hour=timeIn.hour, minute=timeIn.minute)

                seconds = (date1 - date2).total_seconds()
                interval = seconds / 3600

                generalSheet.cell(row=i, column=4).value = interval
                generalSheet.cell(row=i, column=5).value = interval + vorigUur

                vorigUur = generalSheet.cell(row=i, column=5).value
                uren.append(vorigUur)

        for n in range(0, len(sheetNames), 1):
            sheet = wb[sheetNames[n]]
            staal_nr = 7-n

            #print('----NEXT ROW----')
            nextRow = 0
            for cellObj in sheet['A']:
                if cellObj.value == None:
                    nextRow = cellObj.row
                    break
            begin = nextRow

            #print('----INVULLEN----')
            for rows in range(begin, 2+len(uren), 1):
                uurInt = int(round(uren[rows-2],0))
                uurAfgerond = round(uren[rows-2],2)
                sheet.cell(row=nextRow, column=1).value = uurAfgerond

                for i in range(staal_nr, data.max_row, 4):
                    if(str(data.cell(row=i, column=1).value) == str(uurInt)):
                        for j in range(2, 13, 1):
                            sheet.cell(row=nextRow, column=j).value = data.cell(row=i, column=j+7).value
                print(nextRow)
                sheet.cell(row=nextRow, column=17).value = 100-100*sheet.cell(row=nextRow, column=4).value/sheet.cell(row=2, column=4).value
                nextRow = nextRow + 1

        #print('Saving...')
        wb.save('__PID_BIFI_NPERT_JW_5BB_updated.xlsx')
        for i in range(0, len(uren),1):
            uren[i] = int(round(uren[i],0))
        return uren