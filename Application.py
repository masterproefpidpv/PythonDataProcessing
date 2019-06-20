import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter import *
from PIL import ImageTk, Image
import glob
import openpyxl
import os
from os import listdir
import AlgemeneInfo
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
#matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib.figure import Figure
import Data
import AlgemeneInfo
import Main

class Application(tk.Frame):

    def __init__(self, master=None):
        """
        initialisation of the gui
        """
        super().__init__(master)
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        """
        create the different widgets of the gui and put them in rows and columns
        """
        self.grid()
        self.pack(fill="both", expand=True)

        #styles
        bgColor = '#343434' #29648A'
        fgColor = '#FFFFFF' #DFDFDF'
        red = '#f91919'
        gui_style = ttk.Style()
        gui_style.map("My.TButton",
    #foreground=[('pressed', 'red'), ('active', bgColor)],
    background=[('pressed', '!disabled', bgColor), ('active', bgColor), ('!active', bgColor)]
    )
        gui_style.configure('My.TFrame', background=bgColor)
        gui_style.configure('My.TNotebook', background=fgColor)
        gui_style.configure('My.TLabel', background=bgColor, foreground=fgColor, font=('verdana', 10, ''))
        gui_style.configure('My.TCheckbutton', background=bgColor, foreground=fgColor, font=('verdana', 10, ''))
        gui_style.configure('My.TLabelframe.Label', background=bgColor, foreground=fgColor)
        gui_style.configure('My.TRadiobutton', background=bgColor, foreground=fgColor, font=('verdana', 10, ''))
        gui_style.configure('My.TSeparator', background='#000000')

        #tabs
        tabControl = ttk.Notebook(self, style='My.TNotebook')
        tabControl.pack(expand=1, fill='both')

        self.tabBifi = ttk.Frame(tabControl, style='My.TFrame')
        tabControl.add(self.tabBifi, text='Processing')

        self.tabInfo = ttk.Frame(tabControl, style='My.TFrame')
        tabControl.add(self.tabInfo, text='Graphs')

        # self.tabEl = ttk.Frame(tabControl, style='My.TFrame')
        # tabControl.add(self.tabEl, text='EL')

        #variables
        self.f = Figure(figsize=(5, 4), dpi=100)
        self.ax = self.f.add_subplot(111)
        self.fb = Figure(figsize=(5, 4), dpi=100)
        self.bx = self.fb.add_subplot(111)
        self.subfolders = []
        self.filePath = ''
        self.samples = []
        self.wbName = ''
        self.hours = []
        self.hourRb = tk.IntVar()
        self.machineRb = tk.IntVar()
        self.sampleCb = []
        self.sampleRb = tk.IntVar()
        self.hourVar = tk.IntVar()
        self.hourVar2 = tk.IntVar()
        self.timeCb = tk.IntVar()
        self.timeInt = tk.IntVar()
        self.fileVar = tk.StringVar()
        self.comboVar = tk.StringVar()
        self.comboList = []     # nr + file
        self.configContent = [] # nr + filepath
        self.samples = []       # frame + filepath
        self.ivCb = tk.IntVar()
        self.eqeCb = tk.IntVar()
        self.photoCb = tk.IntVar()

        self.filePsc = tk.StringVar()
        self.comboPsc = tk.StringVar()
        self.comboListPsc = ['']
        self.configPsc = []
        self.samplesPsc = []

        self.fileSm = tk.StringVar()
        self.comboSm = tk.StringVar()
        self.comboListSm = [''] # nr + configContent
        self.configSm = []      # filepath

        self.rsh = []

        #set variables if config file excists
        # try:
        #     self.readFile()
        # except IOError:
        #     self.comboList.append('') # voor Psc en Sm ook -----------------------------
        #     pass

        #Processing
            # Machine choise
        ttk.Separator(self.tabBifi, orient=HORIZONTAL, style='My.TSeparator').grid(row=0, columnspan=10, sticky='WE', padx=5)
        self.machineLabel = ttk.Label(self.tabBifi, text='Select device', style='My.TLabel').grid(sticky='W', row=0, column=0, columnspan=2, padx=10)
        self.RbLoana = ttk.Radiobutton(self.tabBifi, text='LOANA', variable=self.machineRb, value=1, style='My.TRadiobutton', command = lambda:self.selectMachine())\
            .grid(row=1,sticky='W', padx=5, pady=(0,5))
        self.RbThin = ttk.Radiobutton(self.tabBifi, text='PME', variable=self.machineRb, value=2, style='My.TRadiobutton', command=lambda: self.selectMachine())\
            .grid(row=1, column=1,sticky='W', pady=(0,5))
        self.machineRb.set(1)
        self.RbSwitch = ttk.Radiobutton(self.tabBifi, text='K2400', variable=self.machineRb, value=3,
                                        style='My.TRadiobutton', command=lambda: self.selectMachine()) \
            .grid(row=1, column=2, sticky='W', pady=(0,5))
        self.machineRb.set(1)
        ttk.Separator(self.tabBifi, orient=HORIZONTAL, style='My.TSeparator').grid(row=2, columnspan=10, sticky='WE', padx=5)
            # Hour choise
        self.hourLabel = ttk.Label(self.tabBifi, text='Stressing duration?', style='My.TLabel').grid(sticky='W', row=2, column=0, columnspan=3, padx=10)
        self.hourRbAll = ttk.Radiobutton(self.tabBifi, text='All', variable=self.hourRb, value=1,style='My.TRadiobutton').grid(row=3, sticky='W', padx=5)
        self.hourRbEntry = ttk.Radiobutton(self.tabBifi, text='Between', variable=self.hourRb, value=2,style='My.TRadiobutton').grid(row=3, column=1,sticky='W')
        self.hourRb.set(1)
        self.hourInput = ttk.Entry(self.tabBifi, textvariable=self.hourVar)
        self.hourInput.grid(sticky='WE', row=3, column=2, pady=5,padx=(5,5))
        self.hourLabel = ttk.Label(self.tabBifi, text='and', style='My.TLabel').grid(sticky='W', row=3,column=3,padx=5)
        self.hourInput2 = ttk.Entry(self.tabBifi, textvariable=self.hourVar2)
        self.hourInput2.grid(sticky='WE', row=3, column=4, pady=5, padx=(5, 5))
        self.timeInterval = ttk.Checkbutton(self.tabBifi, text="Data every", variable=self.timeCb, onvalue=1, offvalue=0,style='My.TCheckbutton')
        self.timeInterval.grid(row=3, column=5, sticky='W', padx=5, pady=5)
        self.timeInput = ttk.Entry(self.tabBifi, textvariable=self.timeInt)
        self.timeInput.grid(sticky='WE', row=3, column=6, pady=5, padx=(5, 5))
        self.everyTimeLabel = ttk.Label(self.tabBifi, text='h', style='My.TLabel')
        self.everyTimeLabel.grid(sticky='W', row=3, column=7, padx=5)
        ttk.Separator(self.tabBifi, orient=HORIZONTAL, style='My.TSeparator').grid(row=4, columnspan=10, sticky='WE', padx=5)
            # Excel file
        self.excelLabel = ttk.Label(self.tabBifi, text='Select Excel file', style='My.TLabel').grid(sticky='W',row=4,column=0,columnspan=2, padx=10)
        #self.configLabel = ttk.Label(self.tabBifi, text='Select previous configuration', style='My.TLabel').grid(sticky='W',row=1,column=3)
        # self.combo = ttk.Combobox(self.tabBifi, textvariable=self.comboVar, values=self.comboList)
        # self.combo.grid(row=5, column=0, columnspan=3, sticky='WE', padx=5)
        # self.combo.current(len(self.comboList) - 1)
        # self.combo.bind('<<ComboboxSelected>>', self.select)
        # self.rmBtn = ttk.Button(self.tabBifi, text='Remove config', style='My.TButton')
        # self.rmBtn.bind('<ButtonRelease-1>', self.remove)
        # self.rmBtn.grid(sticky='WE', row=5, column=4, padx=5)
        self.fileInput = ttk.Entry(self.tabBifi, textvariable=self.fileVar)
        self.fileInput.grid(row=6, column=1, columnspan=8, sticky='WE', padx=5, pady=5)
        self.fileBtn = ttk.Button(self.tabBifi, text='Select file', style='My.TButton')
        self.fileBtn.bind('<ButtonRelease-1>', self.pickFile)
        self.fileBtn.grid(sticky='WE', row=6, column=0, padx=5, pady=5)
            # Data choise
        ttk.Separator(self.tabBifi, orient=HORIZONTAL, style='My.TSeparator').grid(row=7, columnspan=10, sticky='WE', padx=5)
        self.dataLabel = ttk.Label(self.tabBifi, text='Select data types', style='My.TLabel').grid(sticky='W', row=7, column=0,columnspan=2, padx=10)
        self.iv = ttk.Checkbutton(self.tabBifi, text="IV", variable=self.ivCb, onvalue=1, offvalue=0,style='My.TCheckbutton')
        self.iv.grid(row=8, column=0, sticky='W', padx=5, pady=5)
        self.eqe = ttk.Checkbutton(self.tabBifi, text="EQE", variable=self.eqeCb, onvalue=1, offvalue=0,style='My.TCheckbutton')
        self.eqe.grid(row=8, column=1, sticky='W', pady=5)
        self.photo = ttk.Checkbutton(self.tabBifi, text="EL picture", variable=self.photoCb, onvalue=1, offvalue=0,style='My.TCheckbutton')
        self.photo.grid(row=8, column=2, sticky='W', pady=5)
            # Start
        self.beginBtn = ttk.Button(self.tabBifi, text='Begin', command=self.begin, style='My.TButton')
        self.beginBtn.grid(sticky='W',row=9,column=0,padx=5,pady=5)
        self.errorLabel = ttk.Label(self.tabBifi, text='', background=bgColor, foreground=red, font=('verdana', 10, ''))
        self.errorLabel.grid(sticky='W', row=9, column=1, columnspan=3)
        img = Image.open("UHasselt.png")
        basewidth = 120 #70 voor zelfde breedte
        wpercent = (basewidth / float(img.size[0]))
        hsize = int((float(img.size[1]) * float(wpercent)))
        img = img.resize((basewidth, hsize), Image.ANTIALIAS)
        img = ImageTk.PhotoImage(img)
        label1 = Label(self.tabBifi, image=img, background='white')
        label1.image = img
        label1.grid(row=0, column=6, columnspan=2, padx=5, pady=(5,5), sticky='E')

        #configure column 2 to stretch with the window
        self.tabBifi.grid_columnconfigure(7, weight=1)

        #Info
        self.cbFrame = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.cbFrame.grid(row=0, columnspan=5, sticky='W', padx=5)
        self.btn = ttk.Button(self.tabInfo, text='Begin', command=self.graph, style='My.TButton')
        self.btn.grid(sticky='W', row=1, column=0, padx=5, pady=5)
        self.gFrame = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.gFrame.grid(row=2, rowspan=2, sticky='W', padx=5)
        self.gFrame2 = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.gFrame2.grid(row=2, rowspan=2, column=1, sticky='W', padx=5)
        self.tFrame = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.tFrame.grid(row=2, column=2, sticky='W', padx=5)

    # def plotGraph(self):
        self.canvas = FigureCanvasTkAgg(self.f, self.gFrame)
        self.canvas.show()
        self.canvas.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
        self.toolbar = NavigationToolbar2TkAgg(self.canvas, self.gFrame)
        self.toolbar.update()
        self.canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        # Shrink current axis by 25%
        box = self.ax.get_position()
        self.ax.set_position([box.x0, box.y0, box.width * 0.75, box.height])

        self.canvas2 = FigureCanvasTkAgg(self.fb, self.gFrame2)
        self.canvas2.show()
        self.canvas2.get_tk_widget().pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
        self.toolbar = NavigationToolbar2TkAgg(self.canvas2, self.gFrame2)
        self.toolbar.update()
        self.canvas2._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        # Shrink current axis by 25%
        box = self.bx.get_position()
        self.bx.set_position([box.x0, box.y0, box.width * 0.75, box.height])

        # # EL
        # self.pFrame = ttk.Frame(self.tabEl, style='My.TFrame')
        # self.pFrame.grid(row=0, column=0, sticky='W', padx=5)
        # self.tFrameEl = ttk.Frame(self.tabEl, style='My.TFrame')
        # self.tFrameEl.grid(row=1, column=0, sticky='W', padx=5)

    def graph(self):
        # for widget in self.gFrame.winfo_children():
        #     widget.destroy()
        # self.plotGraph()
        self.ax.clear()
        self.bx.clear()
        iv = []
        # i = self.sampleRb.get()
        self.f = plt.figure(1)
        if self.machineRb.get() == 1:
            for i in range(0, len(self.samples)):
                if self.sampleCb[i].get() == 1:
                    for h in range(0, len(self.subfolders)):
                        if os.path.exists(self.filePath + '/' + str(self.subfolders[h]) + '/' + self.samples[i] + '/IV/' +
                                          self.samples[i] + '.lgt'):
                            iv = Data.Data.getDataList(
                                self.filePath + '/' + str(self.subfolders[h]) + '/' + self.samples[i] + '/IV/' +
                                self.samples[i] + '.lgt')
                            div = Data.Data.getDataList(
                                self.filePath + '/' + str(self.subfolders[h]) + '/' + self.samples[i] + '/IV/' +
                                self.samples[i] + '.drk')
                            self.ax.plot(iv[1], iv[0], label=str(self.samples[i] + ' - ' + str(self.hours[h])) + 'h')
                            self.bx.plot(div[1], div[0], label=str(self.samples[i] + ' - ' + str(self.hours[h])) + 'h')

                            self.ax.set_xlabel('Voltage [V]')
                            self.ax.set_ylabel('Current [A]')
                            self.ax.set_title('Light IV')
                            # Put a legend to the right of the current axis
                            self.ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))

                            # self.ax.legend()
                            self.bx.set_xlabel('Voltage [V]')
                            self.bx.set_ylabel('Current [A]')
                            self.bx.set_title('Dark IV')
                            self.bx.legend()
                            # Put a legend to the right of the current axis
                            self.bx.legend(loc='center left', bbox_to_anchor=(1, 0.5))

                # if os.path.exists(self.filePath + '/' + str(self.subfolders[h]) + '/' + self.samples[i] + '/IV/' + self.samples[i] + '.lgt'):
                #     iv = Data.Data.getDataList(self.filePath + '/' + str(self.subfolders[h]) + '/' + self.samples[i] + '/IV/' + self.samples[i] + '.lgt')
                #     div = Data.Data.getDataList(self.filePath + '/' + str(self.subfolders[h]) + '/' + self.samples[i] + '/IV/' + self.samples[i] + '.drk')
                #     self.ax.plot(iv[1], iv[0], label=str(self.hours[h]) + 'h')
                #     self.bx.plot(div[1], div[0], label=str(self.hours[h]) + 'h')
                #
                #     self.ax.set_xlabel('Voltage [V]')
                #     self.ax.set_ylabel('Current [A]')
                #     self.ax.set_title('Light IV')
                #     self.ax.legend()
                #     self.bx.set_xlabel('Voltage [V]')
                #     self.bx.set_ylabel('Current [A]')
                #     self.bx.set_title('Dark IV')
                #     self.bx.legend()

        elif self.machineRb.get() == 2:
            # get IV from excel
            print('thin film')
        else:
            for i in range(0, len(self.samples)):
                if self.sampleCb[i].get() == 1:
                    for h in self.hours:
                        if os.path.exists(self.filePath + '/' + self.samples[i] + '/' + str(h) + '.csv'):
                            iv = Data.Data.getDataListSm(self.filePath + '/' + self.samples[i] + '/' + str(h) + '.csv')
                            self.bx.plot(iv[1], iv[0], label=str(h) + 'min')
                            self.bx.set_xlabel('Voltage [V]')
                            self.bx.set_ylabel('Current [A]')
                            self.bx.set_title('Dark IV')
                            # Put a legend to the right of the current axis
                            self.bx.legend(loc='center left', bbox_to_anchor=(1, 0.5))

                    if os.path.exists(self.filePath + '/' + self.samples[i] + '/' + 'Rsh.csv'):
                        self.rsh = Data.Data.getDataListSm(self.filePath + '/' + self.samples[i] + '/' + 'Rsh.csv')
                        self.ax.plot(self.rsh[1], self.rsh[0], label=str(self.samples[i]))
                        self.ax.set_xlabel('Time [min]')
                        self.ax.set_ylabel('Rsh [Ohm]')
                        self.ax.set_title('Rsh')
                        # Put a legend to the right of the current axis
                        self.ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))

        self.canvas.draw()
        self.canvas2.draw()

        # self.elImage()
        self.table()

    def table(self):
        for widget in self.tFrame.winfo_children():
            widget.destroy()
        if self.machineRb.get() == 1:
            for i in range(0, len(self.samples)):
                if self.sampleCb[i].get() == 1:
                    self.tFrameSample = ttk.Frame(self.tFrame, style='My.TFrame')
                    self.tFrameSample.pack()

                    dirname = self.filePath
                    filespec = str(self.subfolders[-1]) + '-' + '*' + self.wbName
                    file = glob.glob(os.path.join(dirname,filespec))[0]
                    if os.path.exists(file):
                        self.titleS = ttk.Label(self.tFrameSample, text=self.samples[i], style='My.TLabel').grid(sticky='NW',row=0,column=0,padx=5)
                        # ttk.Separator(self.tFrameSample, orient=HORIZONTAL, style='My.TSeparator').grid(row=1, columnspan=9, sticky='WE', padx=5)
                        row = 2
                        columnH = 0
                        columnP = 2
                        columnI = 4
                        columnV = 6
                        columnF = 8

                        self.titleH = ttk.Label(self.tFrameSample, text='Time [h]', style='My.TLabel').grid(sticky='NW', row=row, column=columnH, padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=1, rowspan=len(self.hours)+2, sticky='NS')
                        self.titleP = ttk.Label(self.tFrameSample, text='%PID [%]', style='My.TLabel').grid(sticky='NW', row=row, column=columnP, padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=3,rowspan=len(self.hours) + 2, sticky='NS')
                        self.titleI = ttk.Label(self.tFrameSample, text='Isc [mA]', style='My.TLabel').grid(sticky='NW', row=row, column=columnI,padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=5,rowspan=len(self.hours) + 2, sticky='NS')
                        self.titleV = ttk.Label(self.tFrameSample, text='Voc [mV]', style='My.TLabel').grid(sticky='NW', row=row, column=columnV,padx=5)
                        ttk.Separator(self.tFrameSample, orient=VERTICAL, style='My.TSeparator').grid(row=row, column=7,rowspan=len(self.hours) + 2,sticky='NS')
                        self.titleF = ttk.Label(self.tFrameSample, text='FF [%]', style='My.TLabel').grid(sticky='NW', row=row, column=columnF,padx=5)

                        l=0
                        for h in range(0, len(self.hours)):
                            # get %PID from Excel
                            wb = openpyxl.load_workbook(file, data_only=True)
                            activeSheet = wb[self.samples[i]]
                            if activeSheet.cell(row=h+2, column=17).value is not None:
                                labelP = round(activeSheet.cell(row=h+2, column=17).value,2)
                            else:
                                labelP = '' # label over label -> update
                            if activeSheet.cell(row=h + 2, column=5).value is not None:
                                labelI = activeSheet.cell(row=h + 2, column=5).value
                            else:
                                labelI = ''
                            if activeSheet.cell(row=h + 2, column=7).value is not None:
                                labelV = activeSheet.cell(row=h + 2, column=7).value
                            else:
                                labelV = ''
                            if activeSheet.cell(row=h + 2, column=8).value is not None:
                                labelF = activeSheet.cell(row=h + 2, column=8).value
                            else:
                                labelF = ''
                            l = h + 2+row
                            ttk.Label(self.tFrameSample, text=str(self.hours[h]), style='My.TLabel').grid(row=l, column=columnH, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelP, style='My.TLabel').grid(row=l, column=columnP, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelI, style='My.TLabel').grid(row=l, column=columnI, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelV, style='My.TLabel').grid(row=l, column=columnV, padx=5, sticky='WE')
                            ttk.Label(self.tFrameSample, text=labelF, style='My.TLabel').grid(row=l, column=columnF, padx=5, sticky='WE')

                        ttk.Separator(self.tFrameSample, orient=HORIZONTAL, style='My.TSeparator').grid(row=l+1,columnspan=9,sticky='WE',padx=5)

        elif self.machineRb.get() == 2:
            print('thin film')
        else:
            ttk.Separator(self.tFrame, orient=HORIZONTAL, style='My.TSeparator').grid(row=1, columnspan=15, sticky='WE', padx=5)
            columnH = 0
            for i in range(0, len(self.hours)):
                ttk.Label(self.tFrame, text=str(self.hours[i]), style='My.TLabel').grid(row=i + 2, column=columnH,
                                                                                        padx=5, sticky='WE')

            self.titleH = ttk.Label(self.tFrame, text='Time [min]', style='My.TLabel').grid(sticky='NW', row=0,column=columnH, padx=5)
            ttk.Separator(self.tFrame, orient=VERTICAL, style='My.TSeparator').grid(row=0, column=1,rowspan=len(self.hours) + 2,sticky='NS')

            for s in range(0, len(self.samples)):
                if self.sampleCb[s].get() == 1:
                    if os.path.exists(self.filePath + '/' + self.samples[s] + '/' + 'Rsh.csv'):
                        self.rsh = Data.Data.getDataListSm(self.filePath + '/' + self.samples[s] + '/' + 'Rsh.csv')

                        columnR = 2+s*2

                        # self.titleH = ttk.Label(self.tFrame, text='Time [min]', style='My.TLabel').grid(sticky='NW', row=0, column=columnH, padx=5)
                        ttk.Separator(self.tFrame, orient=VERTICAL, style='My.TSeparator').grid(row=0, column=columnR+1, rowspan=len(self.hours)+2, sticky='NS')
                        self.titleR = ttk.Label(self.tFrame, text=self.samples[s], style='My.TLabel').grid(sticky='NW', row=0, column=columnR, padx=5)

                        i = 0
                        print(self.rsh)
                        for r in range(0, len(self.rsh[1])):
                            print('-----------')
                            print(str(self.rsh[1][r]))
                            print(str(self.hours[i]))
                            if int(self.rsh[1][r]) == int(self.hours[i]):
                                labelR = self.rsh[0][r]
                                print('hour: ' + str(self.hours[i]) + ' - Rsh: ' + str(labelR))

                                ttk.Label(self.tFrame, text=labelR, style='My.TLabel').grid(row=i+2, column=columnR, padx=5, sticky='WE')
                                i = i + 1
                                if i >= len(self.hours):
                                    break

                        print('done')

    def elImage(self):
        # EL image
        for widget in self.pFrame.winfo_children():
            print(widget)
            widget.destroy()
        name = self.wbName.split('.')[0]
        photoName = name + '.jpg'
        file = self.filePath + str(self.subfolders[-1]) + '-' + photoName
        if os.path.exists(file):
            img = Image.open(file)
            hsize = 400
            hpercent = (hsize/float(img.size[1]))
            basewidth = int((float(img.size[0]) * float(hpercent))) #1120
            if basewidth > 1000:
                basewidth = 1000  # 70 voor zelfde breedte
                wpercent = (basewidth / float(img.size[0]))
                hsize = int((float(img.size[1]) * float(wpercent)))
            img = img.resize((basewidth, hsize), Image.ANTIALIAS)
            img = ImageTk.PhotoImage(img)
            label1 = Label(self.pFrame, image=img, background='white')
            label1.image = img
            label1.grid(row=0, column=0, padx=5, pady=(5, 5),
                        sticky='W')  # , columnspan=2) #pady=(0,5) voor breedte


        # EL table

        for widget in self.tFrameEl.winfo_children():
            widget.destroy()
        dirname = self.filePath
        filespec = str(self.subfolders[-1]) + '-' + '*' + self.wbName
        print(dirname)
        print(filespec)
        file = glob.glob(os.path.join(dirname,filespec))[0]
        if os.path.exists(file):

            self.titleN = ttk.Label(self.tFrameEl, text=name, style='My.TLabel').grid(sticky='NW', row=0,column=0, padx=5)
            # ttk.Separator(self.tFrameEl, orient=VERTICAL, style='My.TSeparator').grid(row=0, column=1,rowspan=len(self.samples)*2,sticky='NS')

            # hour titles
            for h in range(0, len(self.hours)):
                self.titleH = ttk.Label(self.tFrameEl, text=str(self.hours[h]) + ' h', style='My.TLabel').grid(sticky='NW', row=0,column=h+1, padx=5)
                # ttk.Separator(self.tFrame, orient=VERTICAL, style='My.TSeparator').grid(row=0, column=h+2,rowspan=len(self.hours) + 2,sticky='NS')

            # ttk.Separator(self.tFrameEl, orient=HORIZONTAL, style='My.TSeparator').grid(row=1,columnspan=len(self.hours) * 2,sticky='WE', padx=5)

            for n in range(0,len(self.samples)):
                self.titleN = ttk.Label(self.tFrameEl, text=self.samples[n], style='My.TLabel').grid(sticky='NW', row=n+1,column=0, padx=5)


            for n in range(0, len(self.samples)):
                for h in range(0, len(self.hours)):
                    wb = openpyxl.load_workbook(file, data_only=True)
                    activeSheet = wb[self.samples[n]]
                    if activeSheet.cell(row=h + 2, column=17).value is not None:
                        labelP = round(activeSheet.cell(row=h + 2, column=17).value, 2)
                    else:
                        labelP = ''
                    self.titleP = ttk.Label(self.tFrameEl, text=labelP, style='My.TLabel').grid(sticky='NW', row=n+1, column=h+1,padx=5)




    def selectMachine(self):
        if self.machineRb.get() == 1:
            self.iv.config(state=NORMAL)
            self.eqe.config(state=NORMAL)
            self.photo.config(state=NORMAL)
            self.everyTimeLabel.config(text='h')
            self.everyTimeLabel.update()
        else:
            self.iv.config(state=DISABLED)
            self.eqe.config(state=DISABLED)
            self.photo.config(state=DISABLED)
            self.everyTimeLabel.config(text='min')
            self.everyTimeLabel.update()

    # def remove(self, event):
    #     """
    #     removes a configuration form the config.txt file
    #     is called by rmBtn
    #     """
    #     caller = event.widget
    #     frame = self.getFrame(caller)
    #
    #     if frame == 'Bifi':
    #         i = int(self.comboVar.get().split(' - ')[0]) - 1
    #         #filepath = self.comboList[i].split('|')[1]
    #         filepath = self.configContent[i].split('-')[1]
    #         print(filepath)
    #         with open('config.txt','r') as file:
    #             filedata = file.read()
    #         filedata= filedata.replace('Bifi|' + filepath + '\n', '')
    #         with open('config.txt', 'w') as file:
    #             file.write(filedata)
    #     # if frame == 'Psc':
    #     #     i = int(self.comboPsc.get().split(' - ')[0]) - 1
    #     #     with open('config.txt', 'r') as file:
    #     #         filedata = file.read()
    #     #     filedata = filedata.replace('Psc|' + self.configPsc[i] + '\n', '')
    #     #     with open('config.txt', 'w') as file:
    #     #         file.write(filedata)

    def pickFile(self, event):
        """
        select a .xlsx file via the filedialog
        is called by fileBtn
        """
        caller = event.widget
        print(caller)
        self.cbFrame.destroy()
        self.cbFrame = None
        self.cbFrame = ttk.Frame(self.tabInfo, style='My.TFrame')
        self.cbFrame.grid(row=0, columnspan=5, sticky='W', padx=5)
        self.filePath = ''
        self.samples = []
        self.wbName = ''
        self.hours = []
        self.sampleRb = tk.IntVar()

        self.filename = tk.filedialog.askopenfilename(initialdir="/", title="Select file",filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
        if self.filename != None:
                self.fileVar.set(self.filename)
                self.wbName = self.filename.split('/')[-1]
                self.filePath = self.filename.replace(self.wbName, '')

    def getSamples(self):
        for f in listdir(self.filePath):
            if self.machineRb.get() == 1:
                if os.path.isdir(self.filePath + f):
                    for g in listdir(self.filePath + f):
                        if os.path.isdir(self.filePath + f + '/' + g):
                            self.samples.append(g)
                            # rb = tk.IntVar()
                break
            elif self.machineRb.get() == 2:
                print('thin film')
            else:
                if os.path.isdir(self.filePath + f):
                    self.samples.append(f)
                    # rb = tk.IntVar()
        if self.samples != []:
            for i in range(0, len(self.samples)):
                # if self.machineRb.get() == 1:
                cbValue = tk.IntVar()
                self.sampleCb.append(cbValue)
                cb = ttk.Checkbutton(self.cbFrame, text=self.samples[i], variable=self.sampleCb[i], onvalue=1, offvalue=0, style='My.TCheckbutton')
                cb.pack(side='left', fill=None, expand=False, padx=(0,5))
                # else:
                #     rb = ttk.Radiobutton(self.cbFrame, text=self.samples[i], variable=self.sampleRb, value=i,
                #                          style='My.TRadiobutton')
                #     rb.pack(side='left', fill=None, expand=False, padx=(0, 5))

    # def select(self, event):
    #     """
    #     select a configuration with the comboBox and set the configuration variables in the gui
    #     """
    #     if self.machineRb.get() == 1:
    #         print(self.comboVar.get())
    #         nr = int(self.comboVar.get().split(' - ')[0])-1
    #         path = self.configContent[nr]
    #         self.setVars('Bifi', path)
    #     if self.machineRb.get() == 2:
    #         print(self.comboPsc.get())
    #         path = self.comboPsc.get().split(' - ')[1]
    #         self.setVars('Psc', path)
    #     if self.machineRb.get() == 3:
    #         print(self.comboSm.get())
    #         path = self.comboSm.get().split(' - ')[1]
    #         self.setVars('Sm', path)
    #
    # def readFile(self):
    #     """
    #     read the config.txt file and put the configurations in the comboBox and in the gui
    #     """
    #     with open('config.txt', 'r') as f:
    #         data = f.read()
    #         if data != '':
    #             self.sampels = data.split('\n')
    #             for i in range(0, len(self.sampels) - 1, 1):
    #                 frame = self.sampels[i].split('|')[0]
    #                 path = self.sampels[i].split('|')[1]
    #                 nr = len(self.configContent) + 1
    #                 self.configContent.append(str(nr) + '-' + path)
    #                 self.setVars(frame, path)
    #                 filenames = self.fileVar.get().split('/')
    #                 filename = filenames[len(filenames)-1]
    #                 self.comboVar.set(str(nr) + ' - ' + filename)
    #                 self.comboList.append(str(self.comboVar.get()))
    #                 # self.comboList.append(self.comboVar.get())
    #                 # if frame == 'Psc':
    #                 #     self.configPsc.append(path)
    #                 #     self.setVars(frame, path)
    #                 #     self.comboPsc.set(str(len(self.configPsc)) + ' - ' + self.filePsc.get())
    #                 #     self.comboListPsc.append(self.comboPsc.get())
    #         else:
    #             self.comboList.append('')
    #         print(self.comboList)
    #         print(self.comboVar.get())
    #
    #
    # def setVars(self, frame, path):
    #     """
    #     set the variables from the configuration from the comboBox in the gui
    #     :param i: the number of the configuration
    #     """
    #     self.fileVar.set(path)
    #     # if frame == 'Psc':
    #     #     self.filePsc.set(path)
    #     # if frame == 'Sm':
    #     #     self.filePsc.set(path)

    def begin(self):
        rb = self.machineRb.get()
        self.getSamples()
        if rb == 1:
            self.beginBifi()
        elif rb == 2:
            self.beginPsc()
        elif rb == 3:
            self.beginSm()
        else:
            self.errorLabel.config(text='Error with machine selection')

    def beginBifi(self):
        """
        check if all variables are set correctly and starts the beginBifi function of the Main class
        is called by beginBtn
        """
        print('beginBifi')

        # check if all variables are set correctly
        # yes: begin
        # no: show error message
        if self.hourRb.get() == 2 and(self.hourInput.get() == '' or self.hourInput2.get() == ''):
            self.errorLabel.config(text='Error : No hour')
        elif self.hourRb.get() == 2 and (not self.hourInput.get().isdigit() or not self.hourInput2.get().isdigit()):
            self.errorLabel.config(text='Error : Hour must be a number')
        elif self.hourRb.get() == 2 and (int(self.hourInput.get()) >= int(self.hourInput2.get())):
            self.errorLabel.config(text='Error : Hour interval is not correct')
        else:
            # if begin function is done : show 'file saved' and save the configuration if not already in config.txt
            if(self.fileInput.get().endswith('.xlsx')):
                self.errorLabel.config(text='Running...')
                self.errorLabel.update()

                #calculate hours
                #all hours
                self.hours = AlgemeneInfo.AlgemeneInfo.calculateHours(self.wbName, self.filePath)
                self.subfolders = [f.name for f in os.scandir(self.filePath) if f.is_dir()]
                self.subfolders.sort(key=float)
                #between hours
                if self.hourRb.get() != 1:
                    times = self.hours
                    self.hours = []
                    allsubfolders = self.subfolders
                    self.subfolders = []
                    for h in range(0, len(times)):
                        if int(self.hourInput.get()) <= int(times[h]) and int(times[h]) <= int(self.hourInput2.get()):
                            self.hours.append(int(times[h]))
                            self.subfolders.append(allsubfolders[h])
                #data every
                if self.timeCb.get() == 1:
                    print('---timeCb---')
                    threshold = int(self.timeInt.get())
                    times = self.hours
                    self.hours = [times[0]]
                    allsubfolders = self.subfolders
                    self.subfolders = [allsubfolders[0]]
                    lastVal = 0
                    for h in range(0, len(times)):
                        if int(times[h]) >= int(lastVal) + int(threshold):
                            self.hours.append(times[h])
                            self.subfolders.append(allsubfolders[h])
                            lastVal = int(times[h])

                print(self.hours)
                print(self.subfolders)
                if Main.Main.beginBifi(self, self.hours, self.subfolders, self.wbName, self.filePath, self.ivCb.get(), self.eqeCb.get(), self.photoCb.get()):
                    self.errorLabel.config(text='File saved')

                else:
                    self.errorLabel.config(text='Error: Something went wrong')
            else:
                self.errorLabel.config(text='Error : Not a .xlsx file')

    def beginPsc(self):
        """
        check if all variables are set correctly and starts the beginPsc function of the Main class
        is called by beginBtnPsc
        """
        print('beginPsc')
        # check if all variables are set correctly
        # yes: begin
        # no: show error message
        # if begin function is done : show 'file saved' and save the configuration if not already in config.txt
        if(self.fileInput.get().endswith('.xlsx')):
            self.errorLabel.config(text='Running...')
            self.errorLabel.update()
            if Main.Main.beginPsc(self, self.wbName, self.filePath):
                self.errorLabel.config(text='File saved')
            else:
                self.errorLabel.config(text='Error: Something went wrong')
        else:
            self.errorLabel.config(text='Error : Not a .xlsx file')

    def beginSm(self):
        """
        check if all variables are set correctly and starts the beginSm function of the Main class
        is called by beginBtnSm
        """
        modulePath = self.filePath + self.samples[0]
        files = [f.name for f in os.scandir(modulePath)]
        for f in files:
            if f.endswith('.csv') and not f.startswith('Rsh'):
                self.hours.append(f[:-4])
        self.hours.sort(key=int)

        if self.hourRb.get() == 2:
            times = self.hours
            self.hours = []
            for h in times:
                if int(self.hourInput.get()) <= int(h) and int(h) <= int(self.hourInput2.get()):
                    self.hours.append(int(h))
        print(self.hours)
        if self.timeCb.get() == 1:
            print('---timeCb---')
            threshold = int(self.timeInt.get())
            times = self.hours
            self.hours = [0]
            lastVal = 0
            for h in times:
                if int(h) >= int(lastVal) + int(threshold):
                    self.hours.append(int(h))
                    lastVal = int(h)

        # check if all variables are set correctly
        # yes: begin
        # no: show error message

        # if begin function is done : show 'file saved' and save the configuration if not already in config.txt
        if(self.fileInput.get().endswith('.xlsx')):
            self.errorLabel.config(text='Running...')
            self.errorLabel.update()
            if Main.Main.beginSm(self, self.wbName, self.filePath, self.hours):
                self.errorLabel.config(text='File saved')
            else:
                self.errorLabel.config(text='Error: Something went wrong')
        else:
            self.errorLabel.config(text='Error : Not a .xlsx file')



root = Tk()
root.title("PID data automation")
root.tk.call('wm', 'iconphoto', root._w, PhotoImage(file='Monocrystalline-solar-cell.png'))
app = Application(master=root)
app.mainloop()